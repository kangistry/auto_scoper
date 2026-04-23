"""
Script to add clickable cloud-hosted PDF links to the questions spreadsheet.
Supports GitHub, Google Drive, and OneDrive.

Usage:
    python add_pdf_links_cloud.py --github yourusername/your-repo
    python add_pdf_links_cloud.py --onedrive "https://your-sharepoint-site.com/folder"
    python add_pdf_links_cloud.py --gdrive "https://drive.google.com/drive/folders/FOLDER_ID"
"""

import pandas as pd
import os
import re
import argparse
from pathlib import Path
from urllib.parse import quote

# Configuration
CSV_FILE = "ALL_PAPERS_COMBINED.csv"
OUTPUT_CSV = "ALL_PAPERS_COMBINED_WITH_LINKS.csv"
OUTPUT_EXCEL = "ALL_PAPERS_WITH_GITHUB_LINKS.xlsx"

# PDF folders to scan
PDF_FOLDERS = [
    "8464C1F_question_pdfs",
    "8464C1H_question_pdfs",
    "8464C2F_question_pdfs",
    "8464C2H_question_pdfs",
]

def get_script_dir():
    return Path(__file__).parent.resolve()

def scan_pdf_folders(base_dir):
    """Scan PDF folders and build index."""
    pdf_index = {}
    
    for folder in PDF_FOLDERS:
        folder_path = base_dir / folder
        if not folder_path.exists():
            print(f"Warning: Folder {folder} not found, skipping...")
            continue
        
        paper_code = folder.replace("_question_pdfs", "")
        
        for pdf_file in folder_path.glob("*.pdf"):
            filename = pdf_file.name
            
            match = re.search(r'(?:AQA-\d+C\d[FH]-)?([A-Z]{3}\d{2})(?:-CR)?_Q+(\d+)(?:_(\d+))?\.pdf', filename, re.IGNORECASE)
            
            if match:
                session = match.group(1).upper()
                question_num = int(match.group(2))
                subq_num = int(match.group(3)) if match.group(3) else None
                
                # Store both the relative path and folder info
                relative_path = f"{folder}/{pdf_file.name}"
                key = (paper_code, session, question_num, subq_num)
                pdf_index[key] = relative_path
    
    print(f"Indexed {len(pdf_index)} PDF files")
    return pdf_index

def extract_paper_info(source_qp):
    if pd.isna(source_qp):
        return None, None
    
    match = re.search(r'8464(C\d[FH])', source_qp)
    paper_code = f"8464{match.group(1)}" if match else None
    
    session_match = re.search(r'(NOV|JUN|OCT)(\d{2})', source_qp, re.IGNORECASE)
    session = f"{session_match.group(1).upper()}{session_match.group(2)}" if session_match else None
    
    return paper_code, session

def extract_subquestion_number(subquestion_id):
    if pd.isna(subquestion_id):
        return None
    
    subq_str = str(subquestion_id)
    
    match = re.search(r'Q\d+\.(\d+)', subq_str)
    if match:
        return int(match.group(1))
    
    match = re.search(r'\d+\.(\d+)', subq_str)
    if match:
        return int(match.group(1))
    
    return None

def find_pdf_for_row(row, pdf_index):
    paper_code, session = extract_paper_info(row['source_qp'])
    question_num = row['question_id']
    subq_num = extract_subquestion_number(row['subquestion_id'])
    
    if not paper_code or not session or pd.isna(question_num):
        return None
    
    question_num = int(question_num)
    
    if subq_num:
        key = (paper_code, session, question_num, subq_num)
        if key in pdf_index:
            return pdf_index[key]
    
    key = (paper_code, session, question_num, None)
    if key in pdf_index:
        return pdf_index[key]
    
    return None

def generate_github_url(relative_path, repo, branch="main"):
    """Generate a direct GitHub raw file URL."""
    # URL encode the path (but keep forward slashes)
    encoded_path = quote(relative_path, safe='/')
    return f"https://github.com/{repo}/raw/{branch}/{encoded_path}"

def generate_onedrive_url(relative_path, base_url):
    """Generate OneDrive/SharePoint URL."""
    encoded_path = quote(relative_path, safe='/')
    base_url = base_url.rstrip('/')
    return f"{base_url}/{encoded_path}"

def generate_gdrive_folder_url(relative_path, folder_url):
    """
    For Google Drive, we can't easily generate direct file links without file IDs.
    This returns a placeholder - user needs to use Google Drive API or manual mapping.
    """
    # Google Drive doesn't support path-based URLs, only file IDs
    # Return a note for the user
    return f"[Requires Google Drive file ID mapping for: {relative_path}]"

def save_excel_with_hyperlinks(df, output_path, url_column='question_pdf_link'):
    """Save DataFrame to Excel with clickable hyperlinks."""
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # Find the URL column index
        url_col_idx = headers.index(url_column) + 1 if url_column in headers else None
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Make URL column clickable
                if col_idx == url_col_idx and value and not pd.isna(value) and str(value).startswith('http'):
                    cell.hyperlink = str(value)
                    cell.value = "📄 Open PDF"
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = value if not pd.isna(value) else ""
        
        # Auto-adjust column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(50, max(12, len(str(header)) + 2))
        
        wb.save(output_path)
        print(f"Saved Excel with clickable hyperlinks: {output_path}")
        return True
        
    except ImportError:
        print("Note: Install openpyxl for Excel output with hyperlinks: pip install openpyxl")
        return False

def main():
    parser = argparse.ArgumentParser(description='Add cloud PDF links to spreadsheet')
    parser.add_argument('--github', metavar='USER/REPO', help='GitHub repository (e.g., myuser/past-papers)')
    parser.add_argument('--branch', default='main', help='GitHub branch (default: main)')
    parser.add_argument('--onedrive', metavar='URL', help='OneDrive/SharePoint base folder URL')
    parser.add_argument('--gdrive', metavar='URL', help='Google Drive folder URL (limited support)')
    parser.add_argument('--local', action='store_true', help='Use local file paths only')
    
    args = parser.parse_args()
    
    # Determine which cloud service to use
    if not any([args.github, args.onedrive, args.gdrive, args.local]):
        print("No cloud service specified. Use --github, --onedrive, --gdrive, or --local")
        print("\nExample:")
        print("  python add_pdf_links_cloud.py --github yourusername/past-papers")
        print("\nGenerating with local paths for now...\n")
        args.local = True
    
    base_dir = get_script_dir()
    print(f"Working directory: {base_dir}")
    
    # Read CSV
    df = pd.read_csv(base_dir / CSV_FILE)
    print(f"Loaded {len(df)} rows")
    
    # Scan PDFs
    pdf_index = scan_pdf_folders(base_dir)
    
    # Generate links
    print("Generating links...")
    pdf_links = []
    
    for idx, row in df.iterrows():
        relative_path = find_pdf_for_row(row, pdf_index)
        
        if relative_path:
            if args.github:
                url = generate_github_url(relative_path, args.github, args.branch)
            elif args.onedrive:
                url = generate_onedrive_url(relative_path, args.onedrive)
            elif args.gdrive:
                url = generate_gdrive_folder_url(relative_path, args.gdrive)
            else:
                url = relative_path  # Local path
            pdf_links.append(url)
        else:
            pdf_links.append(None)
    
    df['question_pdf_link'] = pdf_links
    
    matched = sum(1 for link in pdf_links if link)
    print(f"\nMatched: {matched}/{len(df)} rows")
    
    # Save outputs
    csv_path = base_dir / OUTPUT_CSV
    df.to_csv(csv_path, index=False)
    print(f"Saved CSV: {csv_path}")
    
    # Try to save Excel with hyperlinks
    excel_path = base_dir / OUTPUT_EXCEL
    save_excel_with_hyperlinks(df, excel_path)
    
    # Show sample
    print("\nSample links (first 3):")
    for link in pdf_links[:3]:
        if link:
            print(f"  {link[:80]}..." if len(str(link)) > 80 else f"  {link}")

if __name__ == "__main__":
    main()
